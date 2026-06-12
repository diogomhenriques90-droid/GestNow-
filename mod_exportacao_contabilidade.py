"""
GESTNOW v3 — mod_exportacao_contabilidade.py
Exportação para Contabilidade — Eticadata (TOC)
3 formatos: CSV lançamentos SNC, Excel resumo, PDF relatório mensal
"""
import streamlit as st
import pandas as pd
import uuid, io, os, json
from datetime import datetime, date, timedelta
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                Table, TableStyle, HRFlowable, PageBreak)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from core import save_db, inv, load_db

# ─────────────────────────────────────────────────────────────────
# PLANO DE CONTAS SNC — PADRÃO (editável pelo TOC)
# ─────────────────────────────────────────────────────────────────

CONTAS_SNC_PADRAO = {
    # Vendas / Prestações
    "vendas_servicos":      "7211",   # Prestações de serviços — obras
    "clientes_cc":          "211",    # Clientes c/c
    "iva_liquidado_23":     "24321",  # IVA liquidado 23%
    "iva_liquidado_6":      "24322",  # IVA liquidado 6%
    "iva_liquidado_0":      "24323",  # IVA isento
    # Compras / Fornecedores
    "compras_mat":          "311",    # Compras de mercadorias
    "fse_subempreitadas":   "6221",   # Subempreitadas
    "fse_materiais":        "6222",   # Materiais
    "fse_combustivel":      "6251",   # Deslocações — combustível
    "fse_dormidas":         "6252",   # Deslocações — alojamento
    "fornecedores_cc":      "221",    # Fornecedores c/c
    "iva_dedutivel_23":     "24331",  # IVA dedutível 23%
    "iva_dedutivel_6":      "24332",  # IVA dedutível 6%
    # Pessoal
    "sal_base":             "6311",   # Remunerações base
    "sal_diarias":          "6252",   # Ajudas de custo / diárias
    "tsu_empresa":          "6351",   # Encargos SS — empresa
    "sal_pagar":            "2311",   # Remunerações a pagar
    "ss_pagar":             "2451",   # Contribuições SS a pagar
    "irs_pagar":            "2421",   # IRS retido a pagar
    # Imobilizado
    "amort_equip":          "6421",   # Depreciações — equipamento
    "amort_veic":           "6422",   # Depreciações — viaturas
    "amort_soft":           "6423",   # Depreciações — software
    "dep_acum_equip":       "4381",   # Depreciações acumuladas — equip.
    "dep_acum_veic":        "4382",   # Depreciações acumuladas — veic.
    "dep_acum_soft":        "4383",   # Depreciações acumuladas — soft.
    # Caixa / Banco
    "banco":                "1211",   # Depósitos à ordem — Montepio
    "caixa":                "111",    # Caixa
}

DIARIOS_ETICADATA = {
    "VE": "Vendas",
    "CO": "Compras",
    "OR": "Operações Diversas",
    "SA": "Salários",
    "AM": "Amortizações",
    "BA": "Banco",
}


# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────

def _load(nome, cols):
    try:
        return load_db(nome, cols, silent=True)
    except:
        return pd.DataFrame(columns=cols)

def _num(df, col):
    if df.empty or col not in df.columns:
        return 0.0
    return pd.to_numeric(df[col], errors='coerce').fillna(0).sum()

def _get_contas() -> dict:
    """Carrega contas SNC configuradas ou usa padrão."""
    try:
        from core import _gcs_read
        buf = _gcs_read("contabilidade_config.json")
        if buf:
            return json.loads(buf.read().decode('utf-8'))
    except:
        pass
    return CONTAS_SNC_PADRAO.copy()

def _save_contas(contas: dict):
    try:
        from core import _gcs_write_binary
        _gcs_write_binary(
            json.dumps(contas, indent=2, ensure_ascii=False).encode('utf-8'),
            "contabilidade_config.json"
        )
    except:
        pass


# ─────────────────────────────────────────────────────────────────
# MOTOR DE LANÇAMENTOS CONTABILÍSTICOS
# ─────────────────────────────────────────────────────────────────

def _gerar_lancamentos(mes: int, ano: int,
                        fat_cli, fat_forn,
                        rh_db, diarias_pag,
                        imob_db, contas: dict) -> list:
    """
    Gera lista de lançamentos contabilísticos do mês
    no formato Eticadata (uma linha por movimento).
    """
    lancamentos = []
    n_lanc = 0

    def _add(n_ordem, diario, tipo_doc, data_str,
             conta, descricao, debito, credito,
             nif="", doc_origem=""):
        lancamentos.append({
            "N_Ordem_Lancamento": n_ordem,
            "Diario":             diario,
            "Tipo_Documento":     tipo_doc,
            "Data":               data_str,
            "Conta":              conta,
            "Descricao":          descricao[:60],
            "Debito":             round(debito, 2) if debito else "",
            "Credito":            round(credito, 2) if credito else "",
            "NIF":                nif,
            "Doc_Origem":         doc_origem,
        })

    # ── 1. FATURAS A CLIENTES (Diário VE) ─────────────────────
    if not fat_cli.empty and 'Data_Emissao' in fat_cli.columns:
        fc = fat_cli.copy()
        fc['Data_d'] = pd.to_datetime(
            fc['Data_Emissao'], dayfirst=True, errors='coerce'
        )
        fc['Total_N']    = pd.to_numeric(fc.get('Total',0), errors='coerce').fillna(0)
        fc['Subtotal_N'] = pd.to_numeric(fc.get('Subtotal',0), errors='coerce').fillna(0)
        fc['IVA_N']      = pd.to_numeric(fc.get('IVA',0), errors='coerce').fillna(0)

        fc_mes = fc[
            (fc['Data_d'].dt.month == mes) &
            (fc['Data_d'].dt.year  == ano) &
            (~fc.get('Estado','').isin(['Anulada']))
        ]

        for _, fat in fc_mes.iterrows():
            n_lanc += 1
            data_s = fat['Data_d'].strftime('%d/%m/%Y') \
                     if pd.notna(fat['Data_d']) else ""
            num    = str(fat.get('Numero',''))
            cli    = str(fat.get('Cliente',''))[:40]
            nif_c  = str(fat.get('NIF_Cliente',''))
            sub    = float(fat.get('Subtotal_N',0))
            iva    = float(fat.get('IVA_N',0))
            tot    = float(fat.get('Total_N',0))
            pct_iva= round(iva/sub*100) if sub > 0 else 23
            conta_iva = (
                contas['iva_liquidado_23'] if pct_iva >= 23 else
                contas['iva_liquidado_6']  if pct_iva >= 6  else
                contas['iva_liquidado_0']
            )

            desc = f"Fatura {num} — {cli}"

            # Débito: Cliente
            _add(n_lanc,"VE","FT",data_s,
                 contas['clientes_cc'],desc,tot,0,nif_c,num)
            # Crédito: Prestação de serviços
            _add(n_lanc,"VE","FT",data_s,
                 contas['vendas_servicos'],desc,0,sub,nif_c,num)
            # Crédito: IVA liquidado
            if iva > 0:
                _add(n_lanc,"VE","FT",data_s,
                     conta_iva,f"IVA {desc}",0,iva,nif_c,num)

    # ── 2. FATURAS DE FORNECEDORES (Diário CO) ────────────────
    if not fat_forn.empty and 'Data' in fat_forn.columns:
        ff = fat_forn.copy()
        ff['Data_d'] = pd.to_datetime(
            ff['Data'], dayfirst=True, errors='coerce'
        )
        ff['Total_N']    = pd.to_numeric(ff.get('Total',0), errors='coerce').fillna(0)
        ff['IVA_N']      = pd.to_numeric(ff.get('IVA',0), errors='coerce').fillna(0)
        ff['Ret_N']      = pd.to_numeric(ff.get('Retencao_Val',0), errors='coerce').fillna(0)

        ff_mes = ff[
            (ff['Data_d'].dt.month == mes) &
            (ff['Data_d'].dt.year  == ano)
        ]

        for _, fat in ff_mes.iterrows():
            n_lanc += 1
            data_s   = fat['Data_d'].strftime('%d/%m/%Y') \
                       if pd.notna(fat['Data_d']) else ""
            forn     = str(fat.get('Fornecedor',''))[:40]
            num_f    = str(fat.get('Numero_Fatura',''))
            tot_f    = float(fat.get('Total_N',0))
            iva_f    = float(fat.get('IVA_N',0))
            ret_f    = float(fat.get('Ret_N',0))
            sub_f    = tot_f - iva_f
            desc_f   = str(fat.get('Descricao',''))[:30]

            # Conta FSE por tipo
            tipo_f = str(fat.get('Tipo_Fornecedor',
                                  fat.get('Categoria','FSE'))).lower()
            if 'sub' in tipo_f or 'empreit' in tipo_f:
                conta_fse = contas['fse_subempreitadas']
            elif 'comb' in tipo_f or 'gasoleo' in tipo_f:
                conta_fse = contas['fse_combustivel']
            elif 'dorm' in tipo_f or 'hotel' in tipo_f:
                conta_fse = contas['fse_dormidas']
            else:
                conta_fse = contas['fse_materiais']

            desc = f"Fat. {forn} {num_f} — {desc_f}"

            # Débito: FSE
            _add(n_lanc,"CO","VF",data_s,
                 conta_fse,desc,sub_f,0,"",num_f)
            # Débito: IVA dedutível
            if iva_f > 0:
                _add(n_lanc,"CO","VF",data_s,
                     contas['iva_dedutivel_23'],
                     f"IVA {desc}",iva_f,0,"",num_f)
            # Crédito: Fornecedor (valor líquido de retenção)
            _add(n_lanc,"CO","VF",data_s,
                 contas['fornecedores_cc'],desc,0,
                 tot_f - ret_f,"",num_f)
            # Crédito: Retenção na fonte
            if ret_f > 0:
                _add(n_lanc,"CO","VF",data_s,
                     contas['irs_pagar'],
                     f"Retenção {forn}",0,ret_f,"",num_f)

    # ── 3. SALÁRIOS (Diário SA) ────────────────────────────────
    if not rh_db.empty and 'Salario_Base' in rh_db.columns:
        TSU_EMP  = 0.2375
        TSU_TRAB = 0.11
        IRS_EST  = 0.08   # estimativa média

        data_sal = f"28/{mes:02d}/{ano}"  # data habitual processamento
        n_lanc += 1

        tot_sal   = _num(rh_db,'Salario_Base')
        tot_tsu_e = round(tot_sal * TSU_EMP,  2)
        tot_tsu_t = round(tot_sal * TSU_TRAB, 2)
        tot_irs   = round(tot_sal * IRS_EST,  2)
        tot_liq   = round(tot_sal - tot_tsu_t - tot_irs, 2)

        desc_sal = f"Processamento salários {mes:02d}/{ano}"

        # Débito: Remunerações base
        _add(n_lanc,"SA","PP",data_sal,
             contas['sal_base'],desc_sal,tot_sal,0)
        # Crédito: Remunerações a pagar (líquido)
        _add(n_lanc,"SA","PP",data_sal,
             contas['sal_pagar'],desc_sal,0,tot_liq)
        # Crédito: IRS a pagar
        _add(n_lanc,"SA","PP",data_sal,
             contas['irs_pagar'],f"IRS retido {mes:02d}/{ano}",
             0,tot_irs)
        # Crédito: SS trabalhador
        _add(n_lanc,"SA","PP",data_sal,
             contas['ss_pagar'],f"SS trabalhador {mes:02d}/{ano}",
             0,tot_tsu_t)

        # Encargos SS empresa
        n_lanc += 1
        _add(n_lanc,"SA","PP",data_sal,
             contas['tsu_empresa'],
             f"Encargos SS empresa {mes:02d}/{ano}",
             tot_tsu_e,0)
        _add(n_lanc,"SA","PP",data_sal,
             contas['ss_pagar'],
             f"SS empresa {mes:02d}/{ano}",
             0,tot_tsu_e)

    # ── 4. DIÁRIAS (Diário OR) ────────────────────────────────
    if not diarias_pag.empty and 'Data_Pagamento' in diarias_pag.columns:
        dp = diarias_pag.copy()
        dp['Data_d'] = pd.to_datetime(
            dp['Data_Pagamento'], dayfirst=True, errors='coerce'
        )
        dp['Val_N'] = pd.to_numeric(
            dp.get('Valor_Total',0), errors='coerce'
        ).fillna(0)

        dp_mes = dp[
            (dp['Data_d'].dt.month == mes) &
            (dp['Data_d'].dt.year  == ano) &
            (dp.get('Status','') == 'Pago')
        ]

        if not dp_mes.empty:
            n_lanc += 1
            tot_diar = dp_mes['Val_N'].sum()
            data_diar = f"30/{mes:02d}/{ano}"
            desc_diar = f"Diárias/ajudas custo {mes:02d}/{ano}"

            _add(n_lanc,"OR","LC",data_diar,
                 contas['sal_diarias'],desc_diar,tot_diar,0)
            _add(n_lanc,"OR","LC",data_diar,
                 contas['banco'],desc_diar,0,tot_diar)

    # ── 5. AMORTIZAÇÕES (Diário AM) ───────────────────────────
    if not imob_db.empty and 'Amort_Anual' in imob_db.columns:
        imob_db2 = imob_db.copy()
        imob_db2['Amort_M'] = pd.to_numeric(
            imob_db2['Amort_Anual'], errors='coerce'
        ).fillna(0) / 12

        data_amort = f"30/{mes:02d}/{ano}"

        for _, ativo in imob_db2.iterrows():
            amort_m = float(ativo.get('Amort_M',0))
            if amort_m <= 0:
                continue
            cat = str(ativo.get('Categoria','')).lower()
            if 'viatura' in cat or 'veic' in cat:
                c_gasto  = contas['amort_veic']
                c_dep    = contas['dep_acum_veic']
            elif 'soft' in cat or 'info' in cat:
                c_gasto  = contas['amort_soft']
                c_dep    = contas['dep_acum_soft']
            else:
                c_gasto  = contas['amort_equip']
                c_dep    = contas['dep_acum_equip']

            n_lanc += 1
            desc_am = (
                f"Amort. {str(ativo.get('Descricao',''))[:30]} "
                f"{mes:02d}/{ano}"
            )
            _add(n_lanc,"AM","AM",data_amort,c_gasto,desc_am,amort_m,0)
            _add(n_lanc,"AM","AM",data_amort,c_dep,desc_am,0,amort_m)

    return lancamentos


# ─────────────────────────────────────────────────────────────────
# GERAR EXCEL ETICADATA
# ─────────────────────────────────────────────────────────────────

def _gerar_excel_eticadata(lancamentos: list,
                            mes: int, ano: int) -> bytes:
    """Excel no formato de importação do Eticadata."""
    output = io.BytesIO()

    df_lanc = pd.DataFrame(lancamentos)

    # Renomear para cabeçalhos Eticadata
    df_exp = pd.DataFrame({
        "Número de Ordem Lançamento": df_lanc.get("N_Ordem_Lancamento",""),
        "Diário":                      df_lanc.get("Diario",""),
        "Tipo Documento":              df_lanc.get("Tipo_Documento",""),
        "Data":                        df_lanc.get("Data",""),
        "Conta":                       df_lanc.get("Conta",""),
        "Descrição":                   df_lanc.get("Descricao",""),
        "Débito":                      df_lanc.get("Debito",""),
        "Crédito":                     df_lanc.get("Credito",""),
        "NIF":                         df_lanc.get("NIF",""),
        "Documento Origem":            df_lanc.get("Doc_Origem",""),
    })

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_exp.to_excel(
            writer, index=False,
            sheet_name=f"Lançamentos {mes:02d}-{ano}"
        )
        ws = writer.sheets[f"Lançamentos {mes:02d}-{ano}"]

        # Formatação cabeçalho
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill(
            "solid", fgColor="1E293B"
        )
        header_font = Font(
            color="FFFFFF", bold=True, size=10
        )
        thin = Side(style='thin', color='E2E8F0')
        border = Border(
            left=thin, right=thin, top=thin, bottom=thin
        )

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border    = border

        # Larguras
        col_widths = {
            'A':10,'B':8,'C':10,'D':12,'E':10,
            'F':45,'G':12,'H':12,'I':14,'J':18
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Alternância linhas
        from openpyxl.styles import PatternFill as PF
        fill_alt = PF("solid", fgColor="F8FAFC")
        for i, row in enumerate(ws.iter_rows(
            min_row=2, max_row=ws.max_row
        ), 1):
            if i % 2 == 0:
                for cell in row:
                    cell.fill   = fill_alt
                    cell.border = border
            else:
                for cell in row:
                    cell.border = border

        # Sheet 2 — resumo por diário
        df_res = df_exp.copy()
        df_res['Débito']  = pd.to_numeric(df_res['Débito'], errors='coerce').fillna(0)
        df_res['Crédito'] = pd.to_numeric(df_res['Crédito'],errors='coerce').fillna(0)
        resumo = df_res.groupby('Diário').agg(
            N_Movimentos=('Conta','count'),
            Total_Debito=('Débito','sum'),
            Total_Credito=('Crédito','sum')
        ).reset_index()
        resumo['Diferença'] = resumo['Total_Debito'] - resumo['Total_Credito']
        resumo.to_excel(
            writer, index=False,
            sheet_name="Resumo por Diário"
        )

        # Sheet 3 — resumo por conta
        res_conta = df_res.groupby('Conta').agg(
            Movimentos=('N_Ordem_Lancamento','count'),
            Debito=('Débito','sum'),
            Credito=('Crédito','sum')
        ).reset_index()
        res_conta['Saldo'] = res_conta['Debito'] - res_conta['Credito']
        res_conta = res_conta.sort_values('Conta')
        res_conta.to_excel(
            writer, index=False,
            sheet_name="Balancete Movimento"
        )

    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────
# GERAR CSV ETICADATA
# ─────────────────────────────────────────────────────────────────

def _gerar_csv_eticadata(lancamentos: list) -> bytes:
    """CSV compatível com importação Eticadata."""
    df = pd.DataFrame(lancamentos)
    df_exp = pd.DataFrame({
        "N_Ordem_Lancamento": df.get("N_Ordem_Lancamento",""),
        "Diario":             df.get("Diario",""),
        "Tipo_Documento":     df.get("Tipo_Documento",""),
        "Data":               df.get("Data",""),
        "Conta":              df.get("Conta",""),
        "Descricao":          df.get("Descricao",""),
        "Debito":             df.get("Debito",""),
        "Credito":            df.get("Credito",""),
        "NIF":                df.get("NIF",""),
        "Doc_Origem":         df.get("Doc_Origem",""),
    })
    return df_exp.to_csv(
        index=False, sep=';', encoding='utf-8-sig',
        decimal=','
    ).encode('utf-8-sig')


# ─────────────────────────────────────────────────────────────────
# GERAR PDF RELATÓRIO MENSAL
# ─────────────────────────────────────────────────────────────────

def _gerar_pdf_mensal(mes: int, ano: int,
                       lancamentos: list,
                       resumo_financeiro: dict,
                       empresa: dict) -> bytes:
    meses_pt = ["Janeiro","Fevereiro","Março","Abril","Maio",
                "Junho","Julho","Agosto","Setembro","Outubro",
                "Novembro","Dezembro"]
    mes_nome = meses_pt[mes-1]

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )
    styles = getSampleStyleSheet()
    story  = []

    bold_s = ParagraphStyle(
        'bold', parent=styles['Normal'],
        fontSize=10, fontName='Helvetica-Bold', spaceAfter=3
    )
    sub_s = ParagraphStyle(
        'sub', parent=styles['Normal'],
        fontSize=8, textColor=colors.HexColor('#64748B'), spaceAfter=2
    )
    normal_s = ParagraphStyle(
        'normal', parent=styles['Normal'],
        fontSize=9, spaceAfter=3
    )

    # Cabeçalho
    story.append(Paragraph(
        f"RELATÓRIO MENSAL CONTABILIDADE — {mes_nome.upper()} {ano}",
        ParagraphStyle('titulo', parent=styles['Normal'],
                       fontSize=14, fontName='Helvetica-Bold',
                       textColor=colors.HexColor('#1E293B'),
                       spaceAfter=4)
    ))
    story.append(Paragraph(
        f"{empresa.get('nome','')} · NIF: {empresa.get('nif','')} · "
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        sub_s
    ))
    story.append(HRFlowable(
        width="100%", thickness=2,
        color=colors.HexColor('#3B82F6')
    ))
    story.append(Spacer(1, 0.3*cm))

    # Resumo financeiro
    story.append(Paragraph("<b>RESUMO FINANCEIRO DO MÊS</b>", bold_s))
    res_data = [
        ["Indicador","Valor"],
        ["Faturação Emitida",
         f"€{resumo_financeiro.get('fat_emitida',0):,.2f}"],
        ["Compras / FSE",
         f"€{resumo_financeiro.get('compras',0):,.2f}"],
        ["Salários (brutos)",
         f"€{resumo_financeiro.get('salarios',0):,.2f}"],
        ["Encargos SS empresa",
         f"€{resumo_financeiro.get('tsu_empresa',0):,.2f}"],
        ["Diárias / Ajudas de custo",
         f"€{resumo_financeiro.get('diarias',0):,.2f}"],
        ["Amortizações (estimativa)",
         f"€{resumo_financeiro.get('amortizacoes',0):,.2f}"],
        ["IVA Liquidado",
         f"€{resumo_financeiro.get('iva_liq',0):,.2f}"],
        ["IVA Dedutível",
         f"€{resumo_financeiro.get('iva_ded',0):,.2f}"],
        ["Saldo IVA (a pagar/recuperar)",
         f"€{resumo_financeiro.get('iva_saldo',0):,.2f}"],
    ]
    rt = Table(res_data, colWidths=[10*cm,7*cm])
    rt.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),(-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR',    (0,0),(-1,0), colors.white),
        ('FONTNAME',     (0,0),(-1,0), 'Helvetica-Bold'),
        ('FONTNAME',     (0,-1),(-1,-1),'Helvetica-Bold'),
        ('BACKGROUND',   (0,-1),(-1,-1),colors.HexColor('#EFF6FF')),
        ('FONTSIZE',     (0,0),(-1,-1), 9),
        ('GRID',         (0,0),(-1,-1), 0.3, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS',(0,1),(-1,-2),
         [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING',   (0,0),(-1,-1), 5),
        ('BOTTOMPADDING',(0,0),(-1,-1), 5),
        ('LEFTPADDING',  (0,0),(-1,-1), 6),
        ('ALIGN',        (1,0),(-1,-1), 'RIGHT'),
    ]))
    story.append(rt)
    story.append(Spacer(1, 0.4*cm))

    # Resumo por diário
    story.append(Paragraph("<b>LANÇAMENTOS POR DIÁRIO</b>", bold_s))

    df_l = pd.DataFrame(lancamentos)
    if not df_l.empty:
        df_l['Debito_N']  = pd.to_numeric(df_l.get('Debito',''),  errors='coerce').fillna(0)
        df_l['Credito_N'] = pd.to_numeric(df_l.get('Credito',''), errors='coerce').fillna(0)

        diario_header = [["Diário","Descrição",
                          "Nº Movimentos","Total Débito","Total Crédito"]]
        diario_rows   = []
        tot_d = tot_c = 0.0

        for diario, grupo in df_l.groupby('Diario'):
            n_mov = len(grupo)
            td    = grupo['Debito_N'].sum()
            tc    = grupo['Credito_N'].sum()
            tot_d += td
            tot_c += tc
            diario_rows.append([
                diario,
                DIARIOS_ETICADATA.get(diario, diario),
                str(n_mov),
                f"€{td:,.2f}",
                f"€{tc:,.2f}",
            ])

        diario_rows.append([
            "TOTAL","","",
            f"€{tot_d:,.2f}",
            f"€{tot_c:,.2f}",
        ])

        dt = Table(
            diario_header + diario_rows,
            colWidths=[2*cm,4.5*cm,3.5*cm,4*cm,4*cm]
        )
        dt.setStyle(TableStyle([
            ('BACKGROUND',   (0,0),(-1,0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR',    (0,0),(-1,0), colors.white),
            ('FONTNAME',     (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTNAME',     (0,-1),(-1,-1),'Helvetica-Bold'),
            ('BACKGROUND',   (0,-1),(-1,-1),colors.HexColor('#EFF6FF')),
            ('FONTSIZE',     (0,0),(-1,-1), 8.5),
            ('GRID',         (0,0),(-1,-1), 0.3, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS',(0,1),(-1,-2),
             [colors.white, colors.HexColor('#F8FAFC')]),
            ('TOPPADDING',   (0,0),(-1,-1), 5),
            ('BOTTOMPADDING',(0,0),(-1,-1), 5),
            ('LEFTPADDING',  (0,0),(-1,-1), 6),
            ('ALIGN',        (2,0),(-1,-1), 'RIGHT'),
        ]))
        story.append(dt)

        # Verificação débito = crédito
        story.append(Spacer(1, 0.3*cm))
        ok_equil = abs(tot_d - tot_c) < 0.01
        cor_eq   = '#10B981' if ok_equil else '#EF4444'
        msg_eq   = (
            "✅ EQUILIBRADO — Total Débito = Total Crédito"
            if ok_equil else
            f"⚠️ DESEQUILIBRADO — Diferença: €{abs(tot_d-tot_c):,.2f}"
        )
        story.append(Paragraph(
            f"<font color='{cor_eq}'><b>{msg_eq}</b></font>",
            normal_s
        ))

    story.append(PageBreak())

    # Detalhe de lançamentos
    story.append(Paragraph("<b>DETALHE DOS LANÇAMENTOS</b>", bold_s))
    story.append(HRFlowable(
        width="100%", thickness=1,
        color=colors.HexColor('#E2E8F0')
    ))
    story.append(Spacer(1, 0.2*cm))

    if not df_l.empty:
        lanc_header = [["Nº","Diário","Data","Conta",
                        "Descrição","Débito","Crédito"]]
        lanc_rows   = []
        for _, row in df_l.iterrows():
            lanc_rows.append([
                str(row.get('N_Ordem_Lancamento','')),
                str(row.get('Diario','')),
                str(row.get('Data','')),
                str(row.get('Conta','')),
                str(row.get('Descricao',''))[:40],
                f"€{float(row.get('Debito_N',0)):,.2f}"
                if float(row.get('Debito_N',0))>0 else "",
                f"€{float(row.get('Credito_N',0)):,.2f}"
                if float(row.get('Credito_N',0))>0 else "",
            ])

        lt = Table(
            lanc_header + lanc_rows,
            colWidths=[1*cm,1.5*cm,2.2*cm,2*cm,6.3*cm,2.5*cm,2.5*cm]
        )
        lt.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,0),  colors.HexColor('#1E293B')),
            ('TEXTCOLOR',     (0,0),(-1,0),  colors.white),
            ('FONTNAME',      (0,0),(-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0,0),(-1,-1), 7),
            ('GRID',          (0,0),(-1,-1), 0.2, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),
             [colors.white, colors.HexColor('#F8FAFC')]),
            ('TOPPADDING',    (0,0),(-1,-1), 3),
            ('BOTTOMPADDING', (0,0),(-1,-1), 3),
            ('LEFTPADDING',   (0,0),(-1,-1), 3),
            ('ALIGN',         (5,0),(-1,-1), 'RIGHT'),
        ]))
        story.append(lt)

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"GESTNOW v3.0 · Exportação Eticadata · "
        f"{datetime.now().strftime('%d/%m/%Y %H:%M')} · "
        f"Uso exclusivo TOC/Contabilidade — Confidencial",
        ParagraphStyle('foot', parent=styles['Normal'],
                       fontSize=7, textColor=colors.grey,
                       alignment=1)
    ))
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
# RENDER PRINCIPAL
# ─────────────────────────────────────────────────────────────────

def render_exportacao_contabilidade(*_):
    """Módulo de Exportação para Contabilidade — Eticadata."""

    # ── Carregar dados ────────────────────────────────────────────
    fat_cli  = _load("faturas_clientes.csv",[
        "ID","Numero","Tipo","Data_Emissao","Cliente",
        "NIF_Cliente","Subtotal","IVA","Total","Estado"
    ])
    fat_forn = _load("faturas_fornecedores.csv",[
        "ID","Data","Fornecedor","Numero_Fatura","Descricao",
        "Total","IVA","Retencao_Val","Tipo_Fornecedor","Estado"
    ])
    rh_db    = _load("colaboradores_rh.csv",["Nome","Salario_Base"])
    diarias_pag = _load("diarias_pagamentos.csv",[
        "ID","Data_Pagamento","Técnico","Valor_Total","Status","Obras"
    ])
    imob_db  = _load("imobilizado_db.csv",[
        "ID","Descricao","Categoria","Amort_Anual","Estado"
    ])

    try:
        from mod_admin_diarias import _get_config_empresa
        empresa = _get_config_empresa()
    except:
        empresa = {
            "nome":"Correia Plácido e Sousa, Lda.",
            "nif":"517182718"
        }

    contas   = _get_contas()
    hoje     = date.today()
    meses_pt = ["Janeiro","Fevereiro","Março","Abril","Maio",
                "Junho","Julho","Agosto","Setembro","Outubro",
                "Novembro","Dezembro"]

    # ── CSS ───────────────────────────────────────────────────────
    st.markdown("""
    <style>
    .cont-card {
        background:#1E293B; border-radius:10px;
        padding:12px 16px; margin-bottom:8px;
    }
    .conta-row {
        display:flex; justify-content:space-between;
        padding:5px 0; border-bottom:1px solid #0F172A;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Header ────────────────────────────────────────────────────
    st.markdown(
        "<div style='background:linear-gradient(135deg,#1E293B,#0F172A);"
        "padding:20px;border-radius:14px;margin-bottom:16px;"
        "border:1px solid rgba(255,255,255,0.08);'>"
        "<h2 style='color:#F1F5F9;margin:0;font-size:1.4rem;'>"
        "📤 Exportação para Contabilidade — Eticadata</h2>"
        "<p style='color:#64748B;margin:4px 0 0;font-size:0.85rem;'>"
        f"{empresa.get('nome','')} · "
        "3 formatos: CSV lançamentos · Excel · PDF relatório"
        "</p></div>",
        unsafe_allow_html=True
    )

    # ── Sub-tabs ──────────────────────────────────────────────────
    (t_export, t_preview, t_contas, t_historico) = st.tabs([
        "📤 Exportar Mês",
        "👁️ Preview Lançamentos",
        "⚙️ Plano de Contas SNC",
        "📋 Histórico de Exports",
    ])

    # ════════════════════════════════════════════════════════════════
    # TAB — EXPORTAR MÊS
    # ════════════════════════════════════════════════════════════════
    with t_export:
        st.markdown("### 📤 Exportar Mês para Eticadata")

        col_e1, col_e2 = st.columns(2)
        with col_e1:
            mes_exp = st.selectbox(
                "Mês",
                meses_pt,
                index=hoje.month-1,
                key="exp_mes"
            )
        with col_e2:
            ano_exp = st.number_input(
                "Ano", min_value=2020,
                value=hoje.year, key="exp_ano"
            )

        mes_num = meses_pt.index(mes_exp) + 1

        # Preview rápido do que vai ser exportado
        st.markdown("---")
        st.markdown("#### 📊 O que vai ser exportado")

        # Calcular totais do mês
        def _fat_mes_val(df, col_data, col_val):
            if df.empty or col_data not in df.columns:
                return 0.0
            dc = df.copy()
            dc['d'] = pd.to_datetime(
                dc[col_data], dayfirst=True, errors='coerce'
            )
            dc['v'] = pd.to_numeric(dc.get(col_val,0),errors='coerce').fillna(0)
            return dc[
                (dc['d'].dt.month==mes_num)&(dc['d'].dt.year==ano_exp)
            ]['v'].sum()

        fat_emitida  = _fat_mes_val(fat_cli,'Data_Emissao','Total')
        iva_liq      = _fat_mes_val(fat_cli,'Data_Emissao','IVA')
        compras_val  = _fat_mes_val(fat_forn,'Data','Total')
        iva_ded      = _fat_mes_val(fat_forn,'Data','IVA')
        sal_val      = _num(rh_db,'Salario_Base')
        tsu_e_val    = round(sal_val * 0.2375, 2)
        diarias_val  = 0.0
        if not diarias_pag.empty and 'Data_Pagamento' in diarias_pag.columns:
            dp2 = diarias_pag.copy()
            dp2['d'] = pd.to_datetime(
                dp2['Data_Pagamento'],dayfirst=True,errors='coerce'
            )
            dp2['v'] = pd.to_numeric(dp2.get('Valor_Total',0),errors='coerce').fillna(0)
            diarias_val = dp2[
                (dp2['d'].dt.month==mes_num)&
                (dp2['d'].dt.year==ano_exp)&
                (dp2.get('Status','')=='Pago')
            ]['v'].sum()

        amort_val = _num(imob_db,'Amort_Anual') / 12

        resumo_fin = {
            "fat_emitida":  fat_emitida,
            "compras":      compras_val,
            "salarios":     sal_val,
            "tsu_empresa":  tsu_e_val,
            "diarias":      diarias_val,
            "amortizacoes": amort_val,
            "iva_liq":      iva_liq,
            "iva_ded":      iva_ded,
            "iva_saldo":    round(iva_liq - iva_ded, 2),
        }

        linhas_preview = [
            ("🧾 Faturas emitidas a clientes",
             fat_emitida, "#3B82F6"),
            ("📥 Compras / Fornecedores",
             compras_val, "#F59E0B"),
            ("👥 Salários brutos",
             sal_val, "#8B5CF6"),
            ("🏛️ Encargos SS empresa",
             tsu_e_val, "#8B5CF6"),
            ("💶 Diárias pagas",
             diarias_val, "#10B981"),
            ("📉 Amortizações (estimativa mensal)",
             amort_val, "#06B6D4"),
            ("💰 IVA Liquidado",
             iva_liq, "#EF4444"),
            ("💰 IVA Dedutível",
             iva_ded, "#10B981"),
        ]

        cols_prev = st.columns(4)
        for i, (label, val, cor) in enumerate(linhas_preview):
            with cols_prev[i%4]:
                st.markdown(
                    f"<div class='cont-card' "
                    f"style='border-top:2px solid {cor};'>"
                    f"<p style='color:#64748B;font-size:0.7rem;"
                    f"margin:0 0 4px;text-transform:uppercase;'>"
                    f"{label}</p>"
                    f"<b style='color:{cor};"
                    f"font-size:1.1rem;'>"
                    f"€{val:,.2f}</b>"
                    f"</div>",
                    unsafe_allow_html=True
                )

        # Saldo IVA
        iva_saldo = round(iva_liq - iva_ded, 2)
        cor_iva   = "#EF4444" if iva_saldo > 0 else "#10B981"
        st.markdown(
            f"<div style='background:{cor_iva}12;"
            f"border:1px solid {cor_iva};"
            f"border-radius:8px;padding:10px 16px;"
            f"text-align:center;margin:8px 0;'>"
            f"<b style='color:{cor_iva};'>"
            f"Saldo IVA: €{abs(iva_saldo):,.2f} "
            f"{'a entregar à AT' if iva_saldo>0 else 'a recuperar'}"
            f"</b></div>",
            unsafe_allow_html=True
        )

        st.markdown("---")

        # Botão principal — gerar tudo
        if st.button(
            f"⚙️ Gerar Export {mes_exp} {ano_exp}",
            key="btn_gerar_export",
            type="primary",
            use_container_width=True
        ):
            with st.spinner("A gerar lançamentos contabilísticos..."):
                lancamentos = _gerar_lancamentos(
                    mes_num, ano_exp,
                    fat_cli, fat_forn,
                    rh_db, diarias_pag,
                    imob_db, contas
                )

            if not lancamentos:
                st.warning(
                    "⚠️ Sem dados para este mês. "
                    "Verifica se existem faturas, compras ou "
                    "registos de salários."
                )
            else:
                st.session_state['export_lancamentos'] = lancamentos
                st.session_state['export_mes']         = mes_num
                st.session_state['export_ano']         = ano_exp
                st.session_state['export_mes_nome']    = mes_exp
                st.session_state['export_resumo']      = resumo_fin

                with st.spinner("A gerar Excel..."):
                    excel_b = _gerar_excel_eticadata(
                        lancamentos, mes_num, ano_exp
                    )
                with st.spinner("A gerar CSV..."):
                    csv_b   = _gerar_csv_eticadata(lancamentos)
                with st.spinner("A gerar PDF..."):
                    pdf_b   = _gerar_pdf_mensal(
                        mes_num, ano_exp,
                        lancamentos, resumo_fin, empresa
                    )

                st.session_state['export_excel'] = excel_b
                st.session_state['export_csv']   = csv_b
                st.session_state['export_pdf']   = pdf_b
                st.success(
                    f"✅ {len(lancamentos)} linhas de lançamentos "
                    f"geradas para {mes_exp} {ano_exp}!"
                )
                st.rerun()

        # Downloads
        if st.session_state.get('export_lancamentos'):
            mes_n  = st.session_state['export_mes_nome']
            ano_n  = st.session_state['export_ano']
            n_lanc = len(st.session_state['export_lancamentos'])

            st.markdown(
                f"<div style='background:rgba(16,185,129,0.1);"
                f"border:1px solid #10B981;"
                f"border-radius:8px;padding:12px;"
                f"text-align:center;margin-bottom:12px;'>"
                f"<b style='color:#10B981;'>"
                f"✅ {n_lanc} linhas prontas — {mes_n} {ano_n}</b>"
                f"</div>",
                unsafe_allow_html=True
            )

            col_d1, col_d2, col_d3 = st.columns(3)
            with col_d1:
                st.download_button(
                    "📊 Excel Eticadata\n(3 sheets: lançamentos, resumo, balancete)",
                    data=st.session_state['export_excel'],
                    file_name=(
                        f"eticadata_lancamentos_"
                        f"{mes_n.lower()}_{ano_n}.xlsx"
                    ),
                    mime="application/vnd.openxmlformats-officedocument"
                         ".spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                    key="dl_excel_exp"
                )
                st.markdown(
                    "<small style='color:#64748B;'>"
                    "Eticadata → Contabilidade → Configuração "
                    "→ Importação de Lançamentos</small>",
                    unsafe_allow_html=True
                )
            with col_d2:
                st.download_button(
                    "📄 CSV Lançamentos SNC\n(separador ;  decimal ,)",
                    data=st.session_state['export_csv'],
                    file_name=(
                        f"lancamentos_snc_"
                        f"{mes_n.lower()}_{ano_n}.csv"
                    ),
                    mime="text/csv",
                    use_container_width=True,
                    type="primary",
                    key="dl_csv_exp"
                )
                st.markdown(
                    "<small style='color:#64748B;'>"
                    "Alternativa ao Excel — "
                    "mesmo formato, separador ponto e vírgula</small>",
                    unsafe_allow_html=True
                )
            with col_d3:
                st.download_button(
                    "📋 PDF Relatório TOC\n(resumo + detalhe completo)",
                    data=st.session_state['export_pdf'],
                    file_name=(
                        f"relatorio_contabilidade_"
                        f"{mes_n.lower()}_{ano_n}.pdf"
                    ),
                    mime="application/pdf",
                    use_container_width=True,
                    type="primary",
                    key="dl_pdf_exp"
                )
                st.markdown(
                    "<small style='color:#64748B;'>"
                    "Para arquivo e conferência manual pelo TOC</small>",
                    unsafe_allow_html=True
                )

            # Instrução de importação Eticadata
            st.markdown("---")
            with st.expander("📖 Como importar no Eticadata"):
                st.markdown("""
                **Passos para importar os lançamentos no Eticadata:**

                1. Abre o **Eticadata ERP**
                2. Vai a **Contabilidade → Configuração → Importação de Lançamentos**
                3. Clica em **Assistente de Importação**
                4. Seleciona o ficheiro **Excel** descarregado
                5. Confirma o mapeamento das colunas:
                   - `Número de Ordem Lançamento` → Campo obrigatório
                   - `Diário` → Seleciona o diário correto (VE, CO, SA, OR, AM)
                   - `Conta` → Conta SNC
                   - `Débito` / `Crédito` → Valores
                6. Clica em **Importar**
                7. Verifica os lançamentos no **Diário** e **Balancete**

                > ⚠️ **Nota:** Antes de importar, confirma com o TOC que as
                > contas SNC configuradas no GESTNOW correspondem ao plano
                > de contas do teu Eticadata. Podes ajustar em
                > **⚙️ Plano de Contas SNC**.
                """)

    # ════════════════════════════════════════════════════════════════
    # TAB — PREVIEW LANÇAMENTOS
    # ════════════════════════════════════════════════════════════════
    with t_preview:
        st.markdown("### 👁️ Preview dos Lançamentos")

        if not st.session_state.get('export_lancamentos'):
            st.info(
                "📋 Ainda sem export gerado. "
                "Vai ao tab 📤 Exportar Mês e clica em Gerar."
            )
        else:
            lancamentos = st.session_state['export_lancamentos']
            mes_n  = st.session_state['export_mes_nome']
            ano_n  = st.session_state['export_ano']

            df_prev = pd.DataFrame(lancamentos)

            # KPIs
            df_prev['Deb_N'] = pd.to_numeric(
                df_prev.get('Debito',''), errors='coerce'
            ).fillna(0)
            df_prev['Cred_N'] = pd.to_numeric(
                df_prev.get('Credito',''), errors='coerce'
            ).fillna(0)

            tot_deb   = df_prev['Deb_N'].sum()
            tot_cred  = df_prev['Cred_N'].sum()
            equilibrado = abs(tot_deb - tot_cred) < 0.01

            c1,c2,c3,c4 = st.columns(4)
            with c1: st.metric("📋 Linhas", len(df_prev))
            with c2: st.metric(
                "📊 Lançamentos",
                df_prev['N_Ordem_Lancamento'].nunique()
            )
            with c3: st.metric("💸 Total Débito",  f"€{tot_deb:,.2f}")
            with c4: st.metric("💰 Total Crédito", f"€{tot_cred:,.2f}")

            # Verificação equilíbrio
            if equilibrado:
                st.success("✅ Lançamentos equilibrados — Débito = Crédito")
            else:
                st.error(
                    f"⚠️ Desequilíbrio de "
                    f"€{abs(tot_deb-tot_cred):,.2f} — "
                    f"verificar antes de importar no Eticadata!"
                )

            # Filtro por diário
            diarios_disp = ["Todos"] + df_prev['Diario'].unique().tolist()
            diario_filt  = st.selectbox(
                "Filtrar por Diário",
                diarios_disp, key="prev_diario_filt"
            )
            df_show = df_prev.copy()
            if diario_filt != "Todos":
                df_show = df_show[df_show['Diario']==diario_filt]

            # Tabela com formatação
            cols_show = [c for c in [
                'N_Ordem_Lancamento','Diario','Tipo_Documento',
                'Data','Conta','Descricao','Deb_N','Cred_N',
                'NIF','Doc_Origem'
            ] if c in df_show.columns]
            df_show_disp = df_show[cols_show].rename(columns={
                'N_Ordem_Lancamento':'Nº',
                'Tipo_Documento':'Tipo',
                'Deb_N':'Débito €',
                'Cred_N':'Crédito €',
                'Doc_Origem':'Doc. Origem'
            })
            st.dataframe(
                df_show_disp.style.format({
                    'Débito €':  lambda x: f"€{x:,.2f}" if x > 0 else "",
                    'Crédito €': lambda x: f"€{x:,.2f}" if x > 0 else "",
                }),
                use_container_width=True,
                hide_index=True,
                height=400
            )

            # Resumo por diário e conta
            col_r1, col_r2 = st.columns(2)
            with col_r1:
                st.markdown("#### Por Diário")
                res_d = df_prev.groupby('Diario').agg(
                    Movimentos=('Conta','count'),
                    Debito=('Deb_N','sum'),
                    Credito=('Cred_N','sum')
                ).reset_index()
                res_d['Desc'] = res_d['Diario'].map(DIARIOS_ETICADATA)
                st.dataframe(
                    res_d[['Diario','Desc','Movimentos',
                            'Debito','Credito']],
                    use_container_width=True,
                    hide_index=True
                )
            with col_r2:
                st.markdown("#### Por Conta (Top 10)")
                res_c = df_prev.groupby('Conta').agg(
                    Debito=('Deb_N','sum'),
                    Credito=('Cred_N','sum')
                ).reset_index()
                res_c['Saldo'] = res_c['Debito'] - res_c['Credito']
                res_c = res_c.sort_values(
                    'Debito', ascending=False
                ).head(10)
                st.dataframe(
                    res_c, use_container_width=True, hide_index=True
                )

    # ════════════════════════════════════════════════════════════════
    # TAB — PLANO DE CONTAS SNC
    # ════════════════════════════════════════════════════════════════
    with t_contas:
        st.markdown("### ⚙️ Plano de Contas SNC")
        st.info(
            "Configura as contas SNC a usar nos lançamentos. "
            "O TOC deve confirmar que estas contas existem "
            "no plano de contas do Eticadata da empresa."
        )

        grupos = {
            "📈 Vendas & Clientes": [
                ("vendas_servicos",    "Prestações de Serviços (72xx)"),
                ("clientes_cc",        "Clientes c/c (211)"),
                ("iva_liquidado_23",   "IVA Liquidado 23% (24321)"),
                ("iva_liquidado_6",    "IVA Liquidado 6% (24322)"),
                ("iva_liquidado_0",    "IVA Isento / 0% (24323)"),
            ],
            "📥 Compras & Fornecedores": [
                ("fse_subempreitadas", "FSE — Subempreitadas (6221)"),
                ("fse_materiais",      "FSE — Materiais (6222)"),
                ("fse_combustivel",    "FSE — Combustível (6251)"),
                ("fse_dormidas",       "FSE — Alojamento (6252)"),
                ("fornecedores_cc",    "Fornecedores c/c (221)"),
                ("iva_dedutivel_23",   "IVA Dedutível 23% (24331)"),
                ("iva_dedutivel_6",    "IVA Dedutível 6% (24332)"),
            ],
            "👥 Pessoal & Diárias": [
                ("sal_base",           "Remunerações base (6311)"),
                ("sal_diarias",        "Ajudas de custo / Diárias (6252)"),
                ("tsu_empresa",        "Encargos SS empresa (6351)"),
                ("sal_pagar",          "Remunerações a pagar (2311)"),
                ("ss_pagar",           "SS a pagar (2451)"),
                ("irs_pagar",          "IRS retido a pagar (2421)"),
            ],
            "📉 Imobilizado & Amortizações": [
                ("amort_equip",        "Amortizações — Equipamento (6421)"),
                ("amort_veic",         "Amortizações — Viaturas (6422)"),
                ("amort_soft",         "Amortizações — Software (6423)"),
                ("dep_acum_equip",     "Dep. acumuladas — Equip. (4381)"),
                ("dep_acum_veic",      "Dep. acumuladas — Veic. (4382)"),
                ("dep_acum_soft",      "Dep. acumuladas — Soft. (4383)"),
            ],
            "🏦 Tesouraria": [
                ("banco",              "Banco — Depósitos à ordem (1211)"),
                ("caixa",              "Caixa (111)"),
            ],
        }

        contas_editadas = {}
        with st.form("form_contas_snc"):
            for grupo_nome, campos in grupos.items():
                st.markdown(
                    f"<p style='color:#3B82F6;font-weight:700;"
                    f"font-size:0.88rem;text-transform:uppercase;"
                    f"margin:12px 0 6px;'>{grupo_nome}</p>",
                    unsafe_allow_html=True
                )
                cols_g = st.columns(2)
                for i, (key, label) in enumerate(campos):
                    with cols_g[i%2]:
                        contas_editadas[key] = st.text_input(
                            label,
                            value=contas.get(key, CONTAS_SNC_PADRAO.get(key,'')),
                            key=f"conta_{key}"
                        )

            col_sv1, col_sv2 = st.columns(2)
            with col_sv1:
                if st.form_submit_button(
                    "💾 Guardar Plano de Contas",
                    use_container_width=True, type="primary"
                ):
                    _save_contas(contas_editadas)
                    inv()
                    st.success(
                        "✅ Plano de contas guardado! "
                        "Próximo export usará estas contas."
                    )
                    st.rerun()
            with col_sv2:
                if st.form_submit_button(
                    "🔄 Repor Padrão SNC",
                    use_container_width=True
                ):
                    _save_contas(CONTAS_SNC_PADRAO.copy())
                    inv()
                    st.info("✅ Reposto plano padrão SNC.")
                    st.rerun()

        # Tabela resumo atual
        st.markdown("---")
        st.markdown("#### 📋 Contas Configuradas Atualmente")
        rows_c = []
        for key, label in [
            item for grupo in grupos.values() for item in grupo
        ]:
            rows_c.append({
                "Chave":   key,
                "Descrição":label.split("(")[0].strip(),
                "Conta SNC Atual":contas.get(key,''),
                "Padrão": CONTAS_SNC_PADRAO.get(key,'')
            })
        df_contas = pd.DataFrame(rows_c)
        df_contas['Diferente do Padrão'] = df_contas.apply(
            lambda r: "⚠️ Sim"
            if r['Conta SNC Atual'] != r['Padrão'] else "✅ Padrão",
            axis=1
        )
        st.dataframe(
            df_contas[['Descrição','Conta SNC Atual',
                        'Padrão','Diferente do Padrão']],
            use_container_width=True, hide_index=True
        )

        # Export plano de contas para o TOC
        csv_contas = df_contas.to_csv(
            index=False, encoding='utf-8-sig'
        )
        st.download_button(
            "📥 Exportar Plano de Contas",
            data=csv_contas.encode('utf-8-sig'),
            file_name="plano_contas_snc_gestnow.csv",
            mime="text/csv",
            key="dl_contas"
        )

    # ════════════════════════════════════════════════════════════════
    # TAB — HISTÓRICO DE EXPORTS
    # ════════════════════════════════════════════════════════════════
    with t_historico:
        st.markdown("### 📋 Histórico de Exports")

        # Registar export no histórico
        hist_exp = _load("historico_exports_cont.csv",[
            "ID","Data_Export","Mes","Ano","N_Lancamentos",
            "Total_Debito","Total_Credito","Equilibrado",
            "Exportado_Por"
        ])

        if st.session_state.get('export_lancamentos') and \
           st.session_state.get('export_mes'):
            # Registar automaticamente
            lanc_s = st.session_state['export_lancamentos']
            mes_s  = st.session_state['export_mes']
            ano_s  = st.session_state['export_ano']

            df_s = pd.DataFrame(lanc_s)
            df_s['Deb_N']  = pd.to_numeric(df_s.get('Debito',''),errors='coerce').fillna(0)
            df_s['Cred_N'] = pd.to_numeric(df_s.get('Credito',''),errors='coerce').fillna(0)
            tot_d_s = df_s['Deb_N'].sum()
            tot_c_s = df_s['Cred_N'].sum()

            # Verificar se já está no histórico
            ja_existe = False
            if not hist_exp.empty:
                ja_existe = len(hist_exp[
                    (hist_exp['Mes'].astype(str)==str(mes_s)) &
                    (hist_exp['Ano'].astype(str)==str(ano_s))
                ]) > 0

            if not ja_existe and lanc_s:
                novo_hist = pd.DataFrame([{
                    "ID":            str(uuid.uuid4())[:8].upper(),
                    "Data_Export":   datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "Mes":           mes_s,
                    "Ano":           ano_s,
                    "N_Lancamentos": len(lanc_s),
                    "Total_Debito":  round(tot_d_s,2),
                    "Total_Credito": round(tot_c_s,2),
                    "Equilibrado":   "Sim" if abs(tot_d_s-tot_c_s)<0.01 else "Não",
                    "Exportado_Por": st.session_state.get('user','Admin')
                }])
                upd_h = pd.concat(
                    [hist_exp, novo_hist], ignore_index=True
                ) if not hist_exp.empty else novo_hist
                save_db(upd_h,"historico_exports_cont.csv")
                inv("historico_exports_cont.csv")

        # Recarregar histórico
        hist_exp2 = _load("historico_exports_cont.csv",[
            "ID","Data_Export","Mes","Ano","N_Lancamentos",
            "Total_Debito","Total_Credito","Equilibrado","Exportado_Por"
        ])

        if hist_exp2.empty:
            st.info(
                "📋 Sem exports anteriores. "
                "O histórico é preenchido automaticamente "
                "após cada exportação."
            )
        else:
            meses_pt2 = {str(i+1):m for i,m in enumerate(meses_pt)}
            for _, h in hist_exp2.sort_values(
                'Data_Export', ascending=False
            ).iterrows():
                mes_h  = str(h.get('Mes',''))
                ano_h  = str(h.get('Ano',''))
                eq     = h.get('Equilibrado','')
                cor_eq = "#10B981" if eq=='Sim' else "#EF4444"
                ic_eq  = "✅" if eq=='Sim' else "⚠️"

                st.markdown(
                    f"<div class='cont-card' "
                    f"style='border-left:3px solid {cor_eq};'>"
                    f"<div style='display:flex;"
                    f"justify-content:space-between;'>"
                    f"<div>"
                    f"<b style='color:#F1F5F9;'>"
                    f"📤 {meses_pt2.get(mes_h,mes_h)} {ano_h}</b><br>"
                    f"<small style='color:#64748B;'>"
                    f"Gerado em {h.get('Data_Export','')} · "
                    f"{h.get('N_Lancamentos','')} linhas · "
                    f"Por: {h.get('Exportado_Por','')}"
                    f"</small>"
                    f"</div>"
                    f"<div style='text-align:right;'>"
                    f"<span style='color:{cor_eq};'>"
                    f"{ic_eq} {eq}</span><br>"
                    f"<small style='color:#64748B;'>"
                    f"D: €{float(h.get('Total_Debito',0) or 0):,.2f} · "
                    f"C: €{float(h.get('Total_Credito',0) or 0):,.2f}"
                    f"</small></div></div></div>",
                    unsafe_allow_html=True
                )

            csv_hist = hist_exp2.to_csv(
                index=False, encoding='utf-8-sig'
            )
            st.download_button(
                "📥 Exportar Histórico",
                data=csv_hist.encode('utf-8-sig'),
                file_name="historico_exports_contabilidade.csv",
                mime="text/csv",
                key="dl_hist_exp"
            )
